import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook

class ExcelGUI:
    def __init__(self, master):
        self.master = master
        self.master.title("Involucro 2000")
        self.master.geometry("350x275")

        self.excel_file = "Repuestos2024.xlsx"
        self.create_excel_if_not_exists()

        # Frame para la referencia
        self.ref_frame = tk.Frame(self.master)
        self.ref_frame.pack(pady=10)

        tk.Label(self.ref_frame, text="Ref.Componente").grid(row=0, column=0)
        # Validación para la entrada de referencia
        validate_command = (self.master.register(self.validate_reference), '%P')
        self.ref_entry = tk.Entry(self.ref_frame, validate='key', validatecommand=validate_command)
        self.ref_entry.grid(row=0, column=1)
        # Ahora pulsando ENTER para buscar
        self.ref_entry.bind("<Return>", lambda event: self.buscar_datos())
        tk.Button(self.ref_frame, text="Buscar", command=self.buscar_datos).grid(row=0, column=2)

        # Frame para mostrar y modificar datos
        self.data_frame = tk.Frame(self.master)
        self.data_frame.pack(pady=10)

        self.labels = ["Frontal", "Lateral Der.", "Lateral Izq.", "Power/Reset", "Leds frontales", "Varios", "Protecciones"]
        self.entries = []

        for i, label in enumerate(self.labels):
            tk.Label(self.data_frame, text=label).grid(row=i, column=0)
            if label == "Varios":
                entry = tk.Entry(self.data_frame)  # "Varios" acepta texto
            else:
                validate_command_numeric = (self.master.register(self.validate_numeric), '%P')
                entry = tk.Entry(self.data_frame, validate='key', validatecommand=validate_command_numeric)
            entry.grid(row=i, column=1)
            entry.config(state='disabled')  # Inicialmente desactivado
            entry.bind("<Key>", self.verificar_modificacion)
            self.entries.append(entry)

        # Botón para guardar cambios
        tk.Button(self.master, text="Guardar", command=self.guardar_cambios).pack(pady=10)

    def verificar_modificacion(self, event):
        #Evita que se escriba en el campo sin buscar primero.
        if event.widget['state'] == 'disabled':
            messagebox.showerror("BIEN PERO MAL", "Dale a buscar, paspán")
            return "break"

    def create_excel_if_not_exists(self):
        #Crea el archivo Excel si no existe.
        try:
            load_workbook(self.excel_file)
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["Frontal", "Lateral Der.", "Lateral Izq.", "Power/Reset", "Leds frontales", "Varios", "Protecciones"])
            wb.save(self.excel_file)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear o abrir el archivo Excel: {e}")

    def validate_reference(self, value_if_allowed):
        if value_if_allowed.isdigit() and len(value_if_allowed) <= 5:
            return True
        elif value_if_allowed == "":
            return True  # Permite borrar todo el contenido
        else:
            messagebox.showerror("MAL", "La referencia son 5 dígitos numéricos.")
            return False
    #alida que el campo sea solo numérico
    def validate_numeric(self, value_if_allowed):
        if value_if_allowed.isdigit() or value_if_allowed == "":
            return True
        else:
            messagebox.showerror("MAL PERO BIEN", "Solo números, por favor.")
            return False

    def limpiar_entradas(self, limpiar_referencia=True):
        #Limpia todos los campos de entrada.
        if limpiar_referencia:
            self.ref_entry.delete(0, tk.END)
        for entry in self.entries:
            entry.delete(0, tk.END)

    def buscar_datos(self):
        #Busca los datos de la referencia
        referencia = self.ref_entry.get().strip()
        if not referencia or len(referencia) != 5:
            messagebox.showerror("MAL", "La referencia son 5 dígitos numéricos.")
            return

        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if row[0] == referencia:
                    for entry, value in zip(self.entries, row[1:]):
                        entry.config(state='normal')  # Habilitar entradas al buscar
                        entry.delete(0, tk.END)
                        entry.insert(0, str(value) if value is not None else "")
                    return

            messagebox.showinfo("Nueva Referencia", "No hay repuestos de esta referencia, introduce lo que quieras añadir.")
            self.limpiar_entradas(limpiar_referencia=False)
            for entry in self.entries:
                entry.config(state='normal')

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo buscar en el archivo Excel: {e}")

    def guardar_cambios(self):
        #Guarda los cambios realizados en la referencia actual
        referencia = self.ref_entry.get().strip()
        if not referencia or len(referencia) != 5:
            messagebox.showerror("MAL", "La referencia son 5 dígitos numéricos.")
            return

        datos = [entry.get().strip() for entry in self.entries]

        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active

            row_to_update = None
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[0].value == referencia:
                    row_to_update = row
                    break

            if row_to_update:
                for cell, value in zip(row_to_update[1:], datos):
                    cell.value = value
            else:
                ws.append([referencia] + datos)

            wb.save(self.excel_file)
            messagebox.showinfo("BIEN", "Datos guardados, a currar")
            self.limpiar_entradas(limpiar_referencia=False)
            for entry in self.entries:
                entry.config(state='disabled')  # Desactivar campos después de guardar

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar en el archivo Excel: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelGUI(root)
    root.mainloop()